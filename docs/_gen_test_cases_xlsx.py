# -*- coding: utf-8 -*-
"""从 functional-test-checklist.md 生成测试用例 Excel。

可选合并 docs/functional-test-results.json（执行结果 / 缺陷 / 前置 / 结论）。
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "docs" / "functional-test-checklist.md"
OUT_PATH = ROOT / "docs" / "functional-test-cases.xlsx"
RESULTS_PATH = ROOT / "docs" / "functional-test-results.json"

text = MD_PATH.read_text(encoding="utf-8")
lines = text.splitlines()

run_results: dict = {}
if RESULTS_PATH.exists():
    run_results = json.loads(RESULTS_PATH.read_text(encoding="utf-8"))


def case_run(cid: str) -> dict:
    return (run_results.get("cases") or {}).get(cid) or {}


def smoke_run(sid: str) -> dict:
    return (run_results.get("smoke") or {}).get(sid) or {}

thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
header_fill = PatternFill("solid", fgColor="1F2937")
header_font = Font(color="FFFFFF", bold=True, size=11, name="Microsoft YaHei")
title_font = Font(bold=True, size=16, name="Microsoft YaHei", color="111827")
section_font = Font(bold=True, size=12, name="Microsoft YaHei", color="1F2937")
cell_font = Font(size=10, name="Microsoft YaHei")
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")

p0_fill = PatternFill("solid", fgColor="FEE2E2")
p1_fill = PatternFill("solid", fgColor="FEF3C7")
p2_fill = PatternFill("solid", fgColor="E0E7FF")
pass_fill = PatternFill("solid", fgColor="D1FAE5")
fail_fill = PatternFill("solid", fgColor="FECACA")
block_fill = PatternFill("solid", fgColor="FDE68A")
skip_fill = PatternFill("solid", fgColor="E5E7EB")
smoke_header = PatternFill("solid", fgColor="7C2D12")
info_fill = PatternFill("solid", fgColor="EFF6FF")
amber_fill = PatternFill("solid", fgColor="FEF3C7")
route_fill = PatternFill("solid", fgColor="FFFBEB")


def result_fill(value: str):
    return {
        "通过": pass_fill,
        "失败": fail_fill,
        "阻塞": block_fill,
        "跳过": skip_fill,
    }.get(value)


def is_heading(line: str):
    m = re.match(r"^(#{1,6})\s+(.*)$", line)
    if not m:
        return None
    return len(m.group(1)), m.group(2).strip()


def module_from_h2(h2: str) -> str:
    return re.sub(r"^\d+\.\s*", "", h2)


def parse_tables():
    h2 = ""
    h3 = ""
    i = 0
    n = len(lines)
    while i < n:
        h = is_heading(lines[i])
        if h:
            lv, title = h
            if lv == 2:
                h2 = title
                h3 = ""
            elif lv == 3:
                h3 = title
            i += 1
            continue
        if lines[i].strip().startswith("|") and i + 1 < n and re.match(
            r"^\|\s*[-:]+", lines[i + 1].strip()
        ):
            headers = [c.strip() for c in lines[i].strip().strip("|").split("|")]
            i += 2
            rows = []
            while i < n and lines[i].strip().startswith("|"):
                cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                while len(cells) < len(headers):
                    cells.append("")
                rows.append(cells[: len(headers)])
                i += 1
            yield h2, h3, headers, rows
            continue
        i += 1


def style_header(ws, row: int, cols: int, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_header(ws, row: int = 1):
    ws.freeze_panes = f"A{row + 1}"


def add_result_validation(ws, col_letter: str, start: int, end: int):
    dv = DataValidation(
        type="list",
        formula1='"通过,失败,阻塞,跳过,未测"',
        allow_blank=True,
    )
    dv.error = "请选择列表中的结果"
    dv.errorTitle = "结果"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start}:{col_letter}{end}")


def add_result_cf(ws, col_letter: str, start: int, end: int):
    rng = f"{col_letter}{start}:{col_letter}{end}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"通过"'], fill=pass_fill)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"失败"'], fill=fail_fill)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"阻塞"'], fill=block_fill)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"跳过"'], fill=skip_fill)
    )


cases: list[dict] = []
smoke: list[dict] = []
preflight_items: list[list[str]] = []
route_prep: list[list[str]] = []
known_strategy: list[list[str]] = []
product_confirm: list[list[str]] = []
cadence: list[list[str]] = []

for h2, h3, headers, rows in parse_tables():
    h2_clean = module_from_h2(h2)
    h3_clean = re.sub(r"^\d+(\.\d+)?\s*", "", h3) if h3 else ""

    if h2.startswith("0.") or "测试前置" in h2:
        if headers and headers[0] == "项":
            preflight_items.extend(rows)
        elif headers and "interactionType" in headers[0]:
            route_prep.extend(rows)
        continue

    if h2.startswith("7.") or "Smoke" in h2 or "最小集" in h2:
        for r in rows:
            if len(r) >= 2:
                smoke.append({"id": r[0], "title": r[1]})
        continue

    if h2.startswith("8.") or "临时策略" in h2:
        if headers and "项" in headers and "现状" in "".join(headers):
            known_strategy.extend(rows)
        elif headers and headers[0] == "#":
            product_confirm.extend(rows)
        continue

    if h2.startswith("9.") or "测试节奏" in h2:
        cadence.extend(rows)
        continue

    if h2.startswith("10.") or "结论" in h2:
        continue

    if not headers or headers[0] not in ("#", "编号"):
        continue

    for r in rows:
        cid = r[0].strip()
        if not cid or cid == "#":
            continue

        step = ""
        expected = ""
        title = ""
        extra = ""

        if "题型" in headers:
            idx_type = headers.index("题型")
            idx_step = headers.index("步骤要点") if "步骤要点" in headers else 2
            idx_exp = headers.index("期望") if "期望" in headers else 3
            extra = r[idx_type] if idx_type < len(r) else ""
            step = r[idx_step] if idx_step < len(r) else ""
            expected = r[idx_exp] if idx_exp < len(r) else ""
            title = f"{extra} · {step}" if extra else step
        elif "路径" in headers:
            idx_path = headers.index("路径")
            idx_exp = headers.index("期望") if "期望" in headers else 2
            path = r[idx_path] if idx_path < len(r) else ""
            step = f"访问路径：{path}"
            expected = r[idx_exp] if idx_exp < len(r) else ""
            title = path or cid
        elif "步骤" in headers and "期望" in headers:
            idx_step = headers.index("步骤")
            idx_exp = headers.index("期望")
            step = r[idx_step] if idx_step < len(r) else ""
            expected = r[idx_exp] if idx_exp < len(r) else ""
            title = step[:40] + ("…" if len(step) > 40 else "")
        elif "项" in headers:
            idx_item = headers.index("项")
            step = r[idx_item] if idx_item < len(r) else ""
            expected = step
            title = step[:40]
        else:
            step = " | ".join(r[1:-1]) if len(r) > 2 else (r[1] if len(r) > 1 else "")
            expected = r[-2] if len(r) >= 3 else ""
            title = step[:40]

        major = cid.split(".")[0]
        if cid.startswith("S") or major == "3" or cid.startswith("6.6") or cid.startswith("6.4"):
            prio = "P0"
        elif major in ("1", "4", "6"):
            prio = "P1"
        elif major == "5":
            prio = "P2"
        else:
            prio = "P1"

        pre = "已按「测试前置」准备好环境、账号与对应 interactionType 测试路线；真数据，禁止 mock 主链路。"
        if "后台" in h2 or "web-admin" in h2:
            pre = "已登录 web-admin；具备相应角色权限；H5 环境可用（B→C 相关用例）。"
        if "Ask" in h2 or "问一问" in h2:
            pre = "已登录 H5；网络可访问 Ask/SSE；任务中用例需处于任务上下文。"
        if "鉴权" in h2:
            pre = "H5 可访问；准备合法/非法账号；可清除登录态。"

        cases.append(
            {
                "id": cid,
                "module": h2_clean,
                "submodule": h3_clean,
                "title": title or cid,
                "precondition": pre,
                "steps": step,
                "expected": expected,
                "priority": prio,
                "type": "功能",
                "extra": extra,
            }
        )

# ---------- workbook ----------
wb = Workbook()

# ===== 说明 =====
ws = wb.active
ws.title = "说明"
ws["A1"] = "秘径寻踪 · 功能性测试用例"
ws["A1"].font = title_font
ws.merge_cells("A1:B1")
ws["A2"] = "来源：docs/functional-test-checklist.md"
ws["A3"] = "适用范围：apps/h5-client（主）、apps/web-admin（配套）"
ws["A4"] = "原则：以用户可感知结果验收；真数据优先；禁止用 mock 假任务代替主链路。"
ws["A6"] = "工作表说明"
ws["A6"].font = section_font
ws["A7"] = "工作表"
ws["B7"] = "用途"
style_header(ws, 7, 2)
for i, (a, b) in enumerate(
    [
        ("测试前置", "每次测试必填的环境、账号、路线 ID"),
        ("测试用例", "全量功能用例（步骤、期望、优先级、执行结果）"),
        ("Smoke冒烟", "发版前最小集，约 30~45 分钟"),
        ("缺陷记录", "执行中发现缺陷时填写"),
        ("已知策略", "当前阶段明确策略，勿当缺陷误报"),
        ("测试节奏", "不同场景建议执行范围"),
        ("测试结论", "本次测试汇总签字页"),
    ],
    8,
):
    ws.cell(row=i, column=1, value=a).font = cell_font
    ws.cell(row=i, column=2, value=b).font = cell_font
    for c in range(1, 3):
        ws.cell(row=i, column=c).border = thin
    ws.cell(row=i, column=1).fill = info_fill

ws["A16"] = "严重度建议"
ws["A16"].font = section_font
for r, (code, desc, fill) in enumerate(
    [
        ("P0", "主链路卡死、无法登录、无法通关、数据丢失", p0_fill),
        ("P1", "关键路径错误、进度错误、题型无法提交、Ask 不可用", p1_fill),
        ("P2", "文案/样式/动效、非阻塞体验问题", p2_fill),
    ],
    17,
):
    ws.cell(row=r, column=1, value=code).font = Font(bold=True, size=10, name="Microsoft YaHei")
    ws.cell(row=r, column=2, value=desc).font = cell_font
    for c in range(1, 3):
        ws.cell(row=r, column=c).border = thin
    ws.cell(row=r, column=1).fill = fill

ws["A21"] = "结果填写约定"
ws["A21"].font = section_font
ws["A22"] = "通过 / 失败 / 阻塞 / 跳过 / 未测（下拉选择）"
ws["A24"] = "主链路对照"
ws["A24"].font = section_font
ws["A25"] = (
    "登录 Auth → 展厅 hall → 路线 map → type11 narration 或 "
    "type1~10 brief(locate→video→puzzle/Submit) → result → map 循环 → finale → archive；"
    "配套 Ask SSE"
)
ws["A25"].alignment = wrap
ws.merge_cells("A25:B27")
set_widths(ws, [18, 72])
ws.row_dimensions[1].height = 28
ws.row_dimensions[25].height = 48

# ===== 测试前置 =====
ws = wb.create_sheet("测试前置")
ws["A1"] = "测试前置信息（每次必填）"
ws["A1"].font = title_font
ws.merge_cells("A1:C1")
for c, h in enumerate(["项", "填写内容", "备注"], 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 3)
meta = run_results.get("meta") or {}
preflight_filled = {
    "日期": meta.get("date", ""),
    "测试人": meta.get("tester", ""),
    "前端 commit / 分支": meta.get("branch", "git rev-parse --short HEAD"),
    "后端环境（联调 / 预发 / 生产）": meta.get("backend", "三选一"),
    "后端版本或部署时间": meta.get("backend_version", ""),
    "H5 访问地址": meta.get("h5", ""),
    "后台访问地址": meta.get("admin", ""),
    "测试账号（游客 / 正式）": meta.get("accounts", ""),
    "后台管理员账号": "admin（见 accounts）",
    "后台导游账号": "douzong（见 accounts）",
    "测试路线 ID 列表": meta.get("routes", "type1/6/11 各至少 1 条"),
}
notes = {
    "日期": "YYYY-MM-DD",
    "测试人": "",
    "前端 commit / 分支": "git rev-parse --short HEAD",
    "后端环境（联调 / 预发 / 生产）": "三选一",
    "后端版本或部署时间": "",
    "H5 访问地址": "",
    "后台访问地址": "",
    "测试账号（游客 / 正式）": "",
    "后台管理员账号": "例：admin",
    "后台导游账号": "例：douzong（CREATOR）",
    "测试路线 ID 列表": "type1/6/11 各至少 1 条",
}
for i, row in enumerate(preflight_items, 4):
    item = row[0] if row else ""
    filled = preflight_filled.get(item, row[1] if len(row) > 1 else "")
    ws.cell(row=i, column=1, value=item).font = cell_font
    ws.cell(row=i, column=2, value=filled).font = cell_font
    ws.cell(row=i, column=3, value=notes.get(item, "")).font = cell_font
    for c in range(1, 4):
        ws.cell(row=i, column=c).border = thin
        ws.cell(row=i, column=c).alignment = wrap
    ws.cell(row=i, column=1).fill = info_fill
    if filled:
        ws.cell(row=i, column=2).fill = pass_fill

start = 4 + len(preflight_items) + 1
ws.cell(row=start, column=1, value="建议准备的测试路线").font = section_font
for c, h in enumerate(["interactionType", "用途", "路线 ID", "备注"], 1):
    ws.cell(row=start + 1, column=c, value=h)
style_header(ws, start + 1, 4)
for i, row in enumerate(route_prep, start + 2):
    for c, val in enumerate(row[:4], 1):
        ws.cell(row=i, column=c, value=val).font = cell_font
        ws.cell(row=i, column=c).border = thin
        ws.cell(row=i, column=c).alignment = wrap
    ws.cell(row=i, column=3).fill = route_fill
set_widths(ws, [36, 36, 28, 36])
ws.row_dimensions[1].height = 26

# ===== 测试用例 =====
ws = wb.create_sheet("测试用例")
case_headers = [
    "用例编号",
    "模块",
    "子模块",
    "用例标题",
    "优先级",
    "用例类型",
    "前置条件",
    "测试步骤",
    "预期结果",
    "执行结果",
    "实际结果",
    "缺陷ID",
    "测试人",
    "测试日期",
    "备注",
]
for c, h in enumerate(case_headers, 1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(case_headers))
ws.row_dimensions[1].height = 22
ws.auto_filter.ref = f"A1:{get_column_letter(len(case_headers))}1"
freeze_header(ws, 1)

meta = run_results.get("meta") or {}
tester = meta.get("tester", "")
test_date = meta.get("date", "")
result_counts = {"通过": 0, "失败": 0, "阻塞": 0, "跳过": 0, "未测": 0}

for i, case in enumerate(cases, 2):
    run = case_run(case["id"])
    result = run.get("result") or "未测"
    actual = run.get("actual") or ""
    defect = run.get("defect") or ""
    note = run.get("note") or case.get("extra", "")
    result_counts[result] = result_counts.get(result, 0) + 1
    vals = [
        case["id"],
        case["module"],
        case["submodule"],
        case["title"],
        case["priority"],
        case["type"],
        case["precondition"],
        case["steps"],
        case["expected"],
        result,
        actual,
        defect,
        tester if result != "未测" else "",
        test_date if result != "未测" else "",
        note,
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = cell_font
        cell.border = thin
        cell.alignment = center if c in (1, 5, 6, 10, 13, 14) else wrap
    p = case["priority"]
    if p == "P0":
        ws.cell(row=i, column=5).fill = p0_fill
    elif p == "P1":
        ws.cell(row=i, column=5).fill = p1_fill
    else:
        ws.cell(row=i, column=5).fill = p2_fill
    rf = result_fill(result)
    if rf:
        ws.cell(row=i, column=10).fill = rf
    ws.row_dimensions[i].height = 36

last_case_row = 1 + len(cases)
if cases:
    add_result_validation(ws, "J", 2, last_case_row)
    add_result_cf(ws, "J", 2, last_case_row)

set_widths(ws, [12, 22, 22, 28, 8, 10, 36, 40, 36, 10, 28, 12, 10, 12, 14])

# ===== Smoke =====
ws = wb.create_sheet("Smoke冒烟")
ws["A1"] = "发版前最小集（Smoke，约 30~45 分钟）"
ws["A1"].font = title_font
ws.merge_cells("A1:G1")
ws["A2"] = "时间紧时只跑本表；全量回归再跑「测试用例」全表。"
smoke_headers = ["用例编号", "模块", "测试项", "预期结果", "执行结果", "实际结果", "备注"]
for c, h in enumerate(smoke_headers, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, len(smoke_headers), fill=smoke_header)

if smoke:
    smoke_items = [(s["id"], s["title"]) for s in smoke]
else:
    smoke_items = [
        ("S1", "游客或账号登录成功"),
        ("S2", "展厅列表有已发布路线"),
        ("S3", "任选 1 条 type1 路线：开始 → 跳过识别/播片 → 答题 Submit → 本站结果"),
        ("S4", "任选 1 条 type6：拼图可提交"),
        ("S6", "任选 1 条 type11：解说页完成"),
        ("S7", "刷新后进度可恢复"),
        ("S8", "Ask 发送 1 条，SSE 有回复"),
        ("S9", "通关或至少完成 2 站后 map 状态正确"),
        ("S10", "后台登录 + 打开路线编辑页无报错"),
        ("S11", "（若有内容变更）发布后 H5 可见"),
    ]

for i, (sid, item) in enumerate(smoke_items, 5):
    run = smoke_run(sid)
    result = run.get("result") or "未测"
    actual = run.get("actual") or ""
    note = run.get("defect") or run.get("note") or ""
    vals = [sid, "发版 Smoke", item, item, result, actual, note]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = cell_font
        cell.border = thin
        cell.alignment = center if c in (1, 5) else wrap
    ws.cell(row=i, column=1).fill = p0_fill
    rf = result_fill(result)
    if rf:
        ws.cell(row=i, column=5).fill = rf
    ws.row_dimensions[i].height = 32

last_smoke = 4 + len(smoke_items)
add_result_validation(ws, "E", 5, last_smoke)
add_result_cf(ws, "E", 5, last_smoke)
set_widths(ws, [12, 14, 56, 40, 10, 28, 18])
freeze_header(ws, 4)

# ===== 缺陷记录 =====
ws = wb.create_sheet("缺陷记录")
ws["A1"] = "缺陷记录"
ws["A1"].font = title_font
def_headers = [
    "缺陷ID",
    "模块",
    "关联用例编号",
    "步骤",
    "期望",
    "实际",
    "严重度",
    "截图/录屏",
    "状态",
    "发现人",
    "发现日期",
    "备注",
]
for c, h in enumerate(def_headers, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(def_headers))
defects = run_results.get("defects") or []
for r in range(4, 34):
    d = defects[r - 4] if r - 4 < len(defects) else None
    vals = (
        [
            d.get("id", ""),
            d.get("module", ""),
            d.get("case", ""),
            d.get("steps", ""),
            d.get("expected", ""),
            d.get("actual", ""),
            d.get("severity", ""),
            d.get("screenshot", ""),
            d.get("status", ""),
            tester if d else "",
            test_date if d else "",
            d.get("note", ""),
        ]
        if d
        else [""] * len(def_headers)
    )
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = thin
        cell.font = cell_font
        cell.alignment = wrap
    if d:
        sev = d.get("severity", "")
        if sev == "P0":
            ws.cell(row=r, column=7).fill = p0_fill
        elif sev == "P1":
            ws.cell(row=r, column=7).fill = p1_fill
        elif sev == "P2":
            ws.cell(row=r, column=7).fill = p2_fill
        if d.get("status") == "待修":
            ws.cell(row=r, column=9).fill = fail_fill
    ws.row_dimensions[r].height = 36 if d else 22
dv_sev = DataValidation(type="list", formula1='"P0,P1,P2"', allow_blank=True)
ws.add_data_validation(dv_sev)
dv_sev.add("G4:G33")
dv_st = DataValidation(type="list", formula1='"待修,修复中,已修,延期,不修"', allow_blank=True)
ws.add_data_validation(dv_st)
dv_st.add("I4:I33")
set_widths(ws, [12, 16, 14, 28, 28, 36, 10, 20, 10, 14, 12, 18])

# ===== 已知策略 =====
ws = wb.create_sheet("已知策略")
ws["A1"] = "已知临时策略（验收时勿当缺陷误报）"
ws["A1"].font = title_font
ws.merge_cells("A1:B1")
ws["A2"] = "以下为当前阶段明确策略，除非产品要求收紧，否则不作为 P0 缺陷"
ws["A3"] = "项"
ws["B3"] = "现状"
style_header(ws, 3, 2)
if not known_strategy:
    known_strategy = [
        ["识物 API", "未接公开接口；允许跳过识别，禁止「模拟识别成功」当正式能力"],
        ["播片门槛", "无公开闸门 API；可跳过；无 URL 可用默认片"],
        ["强制跳过本站", "联调/缺接口时本地放行，避免卡死"],
        ["收藏", "通关后偏本地归档，未必同步服务端成就"],
        ["RecordActivity", "弱行为，失败静默"],
        ["Join", "仅用户在 map 主动开始/接着玩"],
    ]
for i, row in enumerate(known_strategy, 4):
    ws.cell(row=i, column=1, value=row[0] if row else "").font = cell_font
    ws.cell(row=i, column=2, value=row[1] if len(row) > 1 else "").font = cell_font
    for c in range(1, 3):
        ws.cell(row=i, column=c).border = thin
        ws.cell(row=i, column=c).alignment = wrap
    ws.cell(row=i, column=1).fill = amber_fill

r0 = 4 + len(known_strategy) + 1
ws.cell(row=r0, column=1, value="产品确认后需改为强制的项（单独跟踪）").font = section_font
for c, h in enumerate(["#", "项", "确认结论", "测试是否改为强制"], 1):
    ws.cell(row=r0 + 1, column=c, value=h)
style_header(ws, r0 + 1, 4)
if not product_confirm:
    product_confirm = [
        ["P1", "识别必须真实成功才能进播片", "", ""],
        ["P2", "生产环境隐藏「跳过本站」", "", ""],
        ["P3", "播片必须真实 URL", "", ""],
    ]
for i, row in enumerate(product_confirm, r0 + 2):
    for c, v in enumerate(row[:4], 1):
        ws.cell(row=i, column=c, value=v).font = cell_font
        ws.cell(row=i, column=c).border = thin
set_widths(ws, [18, 56, 20, 18])

# ===== 测试节奏 =====
ws = wb.create_sheet("测试节奏")
ws["A1"] = "建议测试节奏"
ws["A1"].font = title_font
ws["A3"] = "场景"
ws["B3"] = "跑哪些"
style_header(ws, 3, 2)
if not cadence:
    cadence = [
        ["日常开发自测", "Smoke冒烟表"],
        ["联调日 / 提测", "测试用例中 H5 第1~5章相关 + B→C 闭环（6.6）"],
        ["发版前", "Smoke + 主路径 type1/6/11 全过 + B→C 闭环"],
        ["大改 runtime / adapter", "3.6 进度恢复 + 3.2/3.3 题型"],
        ["大改 Ask", "第 4 章全量"],
        ["大改后台发布流", "6.4 + 6.6"],
    ]
for i, row in enumerate(cadence, 4):
    for c, v in enumerate(row[:2], 1):
        ws.cell(row=i, column=c, value=v).font = cell_font
        ws.cell(row=i, column=c).border = thin
        ws.cell(row=i, column=c).alignment = wrap
    ws.cell(row=i, column=1).fill = info_fill
set_widths(ws, [28, 64])

# ===== 测试结论 =====
ws = wb.create_sheet("测试结论")
ws["A1"] = "本次测试结论（汇总页）"
ws["A1"].font = title_font
ws.merge_cells("A1:B1")
ws["A3"] = "项"
ws["B3"] = "内容"
style_header(ws, 3, 2)
defect_list = run_results.get("defects") or []
executed = sum(result_counts.get(k, 0) for k in ("通过", "失败", "阻塞", "跳过"))
concl_rows = [
    ("总体结论", meta.get("conclusion") or "通过可发版 / 有条件通过 / 不通过"),
    ("P0 缺陷数", str(meta.get("p0", len([d for d in defect_list if d.get("severity") == "P0"])))),
    ("P1 缺陷数", str(meta.get("p1", len([d for d in defect_list if d.get("severity") == "P1"])))),
    ("P2 缺陷数", str(meta.get("p2", len([d for d in defect_list if d.get("severity") == "P2"])))),
    ("阻塞项", meta.get("blockers", "")),
    ("遗留风险", meta.get("risks", "")),
    ("执行用例数（含跳过/阻塞）", str(executed)),
    ("通过数", str(result_counts.get("通过", 0))),
    ("失败数", str(result_counts.get("失败", 0))),
    ("阻塞数", str(result_counts.get("阻塞", 0))),
    ("跳过数", str(result_counts.get("跳过", 0))),
    ("未测数", str(result_counts.get("未测", 0))),
    ("缺陷总数", str(len(defect_list))),
    ("签字", tester),
    ("日期", test_date),
]

for i, (a, b) in enumerate(concl_rows, 4):
    ws.cell(row=i, column=1, value=a).font = cell_font
    ws.cell(row=i, column=2, value=b).font = cell_font
    for c in range(1, 3):
        ws.cell(row=i, column=c).border = thin
        ws.cell(row=i, column=c).alignment = wrap
    ws.cell(row=i, column=1).fill = info_fill
    if a == "总体结论" and b == "有条件通过":
        ws.cell(row=i, column=2).fill = amber_fill
    ws.row_dimensions[i].height = 24

dv_concl = DataValidation(
    type="list", formula1='"通过可发版,有条件通过,不通过"', allow_blank=True
)
ws.add_data_validation(dv_concl)
dv_concl.add("B4")
set_widths(ws, [28, 64])
ws["A21"] = "统计提示"
ws["A21"].font = section_font
ws["A22"] = (
    '结果来源：docs/functional-test-results.json（若存在）合并进本表；'
    '可在「测试用例」按「执行结果」筛选；COUNTIF(测试用例!J:J,"通过")'
)
ws["A22"].font = cell_font
ws.merge_cells("A22:B23")

wb.save(OUT_PATH)
print(f"saved: {OUT_PATH}")
print(f"cases: {len(cases)}, smoke: {len(smoke_items)}, defects: {len(defect_list)}")
print("result counts:", result_counts)
print("sample ids:", [c["id"] for c in cases[:8]], "...", [c["id"] for c in cases[-3:]])
